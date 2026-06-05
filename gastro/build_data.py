#!/usr/bin/env python3
"""Regenerate data.js from the raw research source files.

Run:  python3 build_data.py

This is the single source of truth for how the frontend dataset is built.
Add new doctor reports to EXTRA_REPORTS so no collected data is ever lost.
"""
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent
SRC = Path(
    "/Users/salsabeel/Library/Application Support/Claude/local-agent-mode-sessions/"
    "e9a758f3-8adb-4811-8deb-b255e3792746/b3162796-bd82-480f-8f10-03e4fba681c7/"
    "local_65fe3ac9-fea6-4f6f-b21c-1b711b6fca99/outputs"
)
MAIN_REPORT = SRC / "egypt_gastro_viral_report.xlsx"
ENTITIES_CSV = SRC / "egypt_gastro_entities.csv"
TRANSCRIPTS_CSV = Path(
    "/Users/salsabeel/Downloads/dataset_video-transcriber_2026-06-04_21-03-25-497.csv"
)

# Additional single-doctor reports that use the lightweight (caption-derived) format.
EXTRA_REPORTS = [
    {
        "file": SRC / "egypt_gastro_elsherbiny_report.xlsx",
        "sheet": "El Sherbiny — viral posts",
        "doctor": "Dr. Mohammad El Sherbiny",
        "specialization": "Consultant GIT/Hepatology/Advanced Endoscopy (POEM)",
        "platform": "Instagram",
        "followers": 985,
        "instagram": "https://www.instagram.com/dr.mohammad_elsherbiny/",
        "confidence": "HIGH (from El Sherbiny report)",
        "topic_source": "caption",  # captions are clinical; no audio transcription
    }
]

# Topics that are not clinical "recreate-for-virality" candidates.
PLACEHOLDER = re.compile(r"not in transcribed|not transcribed|no audio|image post", re.I)
NONCLINICAL_EXACT = {
    "Music / no speech",
    "Unclassified",
    "Motivational",
    "Motivational / self-care",
    "Motivational / personal",
    "Personal / family",
    "Professional / ethics",
    "Professional / doctor values",
    "Lifestyle / meds (sponsored)",
}
NONCLINICAL_PREFIX = re.compile(r"^(Conference|Clinic|Personal|Patient appreciation)", re.I)


def norm(s):
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


def toks(s):
    return [t for t in norm(s).split() if t not in {"dr", "doctor"}]


def shortcode(url):
    m = re.search(r"instagram\.com/(?:p|reel|tv)/([^/?#]+)", url or "", re.I)
    return m.group(1) if m else ""


def cap_key(c):
    c = re.sub(r"#\S+", "", c or "")
    c = re.sub(r"[^\u0600-\u06FF a-zA-Z0-9]", " ", c)
    c = re.sub(r"\s+", " ", c).strip().lower()
    return c[:40]


def is_clinical(topic):
    if not topic:
        return False
    if PLACEHOLDER.search(topic) or NONCLINICAL_PREFIX.search(topic):
        return False
    return topic not in NONCLINICAL_EXACT


# ---------- Main viral posts ----------
wb = load_workbook(MAIN_REPORT, read_only=True, data_only=True)
ws = wb["Viral posts"]
rows = list(ws.iter_rows(values_only=True))
ph = [str(x) for x in rows[0]]
pi = {k: i for i, k in enumerate(ph)}
posts = []
rawPosts = []
for r in rows[1:]:
    if not r[pi["Account"]]:
        continue
    rec = {
        "rank": int(r[pi["Rank"]] or 0),
        "account": str(r[pi["Account"]] or "").strip(),
        "platform": str(r[pi["Platform"]] or "").strip(),
        "date": str(r[pi["Date"]] or "").strip(),
        "topicAudio": str(r[pi["Topic (from audio)"]] or "Unclassified").strip(),
        "topicCaption": str(r[pi["Topic (from caption)"]] or "Unclassified").strip(),
        "topicSource": "audio",
        "caption": str(r[pi["Caption"]] or "").strip(),
        "views": int(r[pi["Views"]] or 0),
        "likes": int(r[pi["Likes"]] or 0),
        "comments": int(r[pi["Comments"]] or 0),
        "shares": int(r[pi["Shares"]] or 0),
        "engagement": int(r[pi["Engagement"]] or 0),
        "followers": int(r[pi["Followers"]] or 0),
        "postUrl": str(r[pi["Post URL"]] or "").strip(),
    }
    posts.append(rec)
    rawPosts.append({k: (str(r[pi[k]]) if r[pi[k]] is not None else "") for k in ph})

# ---------- By audio topic sheet (all rows) ----------
topicAudit = []
ws2 = wb["By audio topic"]
rows2 = list(ws2.iter_rows(values_only=True))
for r in rows2[1:]:
    if r[0] is None:
        continue
    topicAudit.append(
        {"topic": str(r[0]).strip(), "posts": int(r[1] or 0), "engagement": int(r[2] or 0)}
    )

# ---------- Entities ----------
entities = []
rawEntities = []
with open(ENTITIES_CSV, newline="", encoding="utf-8") as f:
    for e in csv.DictReader(f):
        rawEntities.append({k: (e.get(k) or "").strip() for k in e})
        entities.append(
            {
                "name": e["Name"].strip(),
                "specialization": e["Specialization"].strip(),
                "platform": e["Platform"].strip(),
                "url": e["URL"].strip(),
                "confidence": e["Confidence"].strip(),
            }
        )

# ---------- Transcripts ----------
transcripts = []
with open(TRANSCRIPTS_CSV, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    fields = reader.fieldnames
    desc_field = fields[0]  # has BOM prefix
    nseg = 0
    while f"transcript/{nseg}/text" in fields:
        nseg += 1
    for row in reader:
        segs = []
        for i in range(nseg):
            t = (row.get(f"transcript/{i}/text") or "").strip()
            if not t:
                continue
            segs.append(
                {
                    "start": float(row.get(f"transcript/{i}/start") or 0),
                    "end": float(row.get(f"transcript/{i}/end") or 0),
                    "text": t,
                }
            )
        full = " ".join(s["text"] for s in segs).strip()
        transcripts.append(
            {
                "videoUrl": (row.get("video_url") or "").strip(),
                "title": (row.get("title") or "").strip(),
                "description": (row.get(desc_field) or "").strip(),
                "duration": float(row.get("duration") or 0),
                "thumbnail": (row.get("thumbnail") or "").strip(),
                "shortcode": shortcode(row.get("video_url") or ""),
                "segments": segs,
                "text": full,
            }
        )

# ---------- Extra single-doctor reports ----------
next_rank = max((p["rank"] for p in posts), default=0)
for cfg in EXTRA_REPORTS:
    ewb = load_workbook(cfg["file"], read_only=True, data_only=True)
    ews = ewb[cfg["sheet"]]
    erows = list(ews.iter_rows(values_only=True))
    ehdr = [str(x) for x in erows[0]]
    ei = {k: i for i, k in enumerate(ehdr)}
    for r in erows[1:]:
        rk = r[ei.get("Rank", 0)]
        if not isinstance(rk, (int, float)):
            continue  # skip metadata/footer rows
        topic = str(r[ei["Topic"]] or "Unclassified").strip()
        next_rank += 1
        rec = {
            "rank": next_rank,
            "account": cfg["doctor"],
            "platform": cfg["platform"],
            "date": str(r[ei["Date"]] or "").strip(),
            "topicAudio": topic,
            "topicCaption": topic,
            "topicSource": cfg["topic_source"],
            "caption": str(r[ei.get("Caption (snippet)", ei.get("Caption", 0))] or "").strip(),
            "views": int(r[ei["Views"]] or 0),
            "likes": int(r[ei["Likes"]] or 0),
            "comments": int(r[ei["Comments"]] or 0),
            "shares": int(r[ei["Shares"]] or 0) if "Shares" in ei else 0,
            "engagement": int(r[ei["Engagement"]] or 0),
            "followers": cfg["followers"],
            "postUrl": str(r[ei["Post URL"]] or "").strip(),
        }
        posts.append(rec)
        rawPosts.append({h: (str(r[ei[h]]) if ei[h] < len(r) and r[ei[h]] is not None else "") for h in ehdr})

    # Register the entity only if it is not already in the entities CSV,
    # so the account stays visible in the Data Vault without duplication.
    already = any(norm(e["name"]) == norm(cfg["doctor"]) for e in entities)
    if not already:
        rawEntities.append(
            {
                "Name": cfg["doctor"],
                "Specialization": cfg["specialization"],
                "Platform": cfg["platform"],
                "URL": cfg["instagram"],
                "Confidence": cfg["confidence"],
            }
        )
        entities.append(
            {
                "name": cfg["doctor"],
                "specialization": cfg["specialization"],
                "platform": cfg["platform"],
                "url": cfg["instagram"],
                "confidence": cfg["confidence"],
            }
        )

# ---------- Join transcripts to posts ----------
tx_by_code = {t["shortcode"]: t for t in transcripts if t["shortcode"]}
ig_capkey_to_tx = {}
for p in posts:
    code = shortcode(p["postUrl"])
    if code and code in tx_by_code:
        tx = tx_by_code[code]
        p["transcript"] = {
            "text": tx["text"],
            "segments": tx["segments"],
            "thumbnail": tx["thumbnail"],
            "duration": tx["duration"],
            "matchedBy": "shortcode",
            "videoUrl": tx["videoUrl"],
        }
        if p["platform"].lower() == "instagram":
            ig_capkey_to_tx[cap_key(p["caption"])] = p["transcript"]
    else:
        p["transcript"] = None
for p in posts:
    if p["transcript"] is None and p["platform"].lower() == "facebook":
        ck = cap_key(p["caption"])
        if ck and ck in ig_capkey_to_tx:
            p["transcript"] = {**ig_capkey_to_tx[ck], "matchedBy": "caption-twin"}

# ---------- Doctors ----------
doctors = {}
by_norm = {}
for e in entities:
    by_norm[norm(e["name"])] = e["name"]
    d = doctors.setdefault(
        e["name"],
        {
            "name": e["name"],
            "specializations": set(),
            "accounts": [],
            "posts": [],
            "image": "",
        },
    )
    d["specializations"].add(e["specialization"])
    d["accounts"].append({"platform": e["platform"], "url": e["url"], "confidence": e["confidence"]})

ent_names = list(doctors.keys())
ent_toks = {n: set(toks(n)) for n in ent_names}


def map_acc(acc):
    nk = norm(acc)
    if nk in by_norm:
        return by_norm[nk]
    at = set(toks(acc))
    best = None
    bs = 0
    for n in ent_names:
        inter = len(at & ent_toks[n])
        if inter > bs:
            bs = inter
            best = n
    return best if best and bs >= 2 else acc


for p in posts:
    p["account"] = map_acc(p["account"])
    d = doctors.setdefault(
        p["account"],
        {
            "name": p["account"],
            "specializations": {"Gastro / Internal Medicine"},
            "accounts": [],
            "posts": [],
            "image": "",
        },
    )
    d["posts"].append(p)

for d in doctors.values():
    existing = {a["platform"].lower() for a in d["accounts"]}
    for plat in ("Facebook", "Instagram"):
        if plat.lower() in existing:
            continue
        for p in d["posts"]:
            if p["platform"].lower() == plat.lower():
                d["accounts"].append(
                    {"platform": plat, "url": p["postUrl"], "confidence": "From viral report"}
                )
                existing.add(plat.lower())
                break
    igp = sorted(
        [p for p in d["posts"] if p["platform"].lower() == "instagram"],
        key=lambda x: (x["engagement"], x["views"]),
        reverse=True,
    )
    for p in igp:
        if p.get("transcript") and p["transcript"].get("thumbnail"):
            d["image"] = p["transcript"]["thumbnail"]
            break

doctorList = []
for d in doctors.values():
    pl = d["posts"]
    doctorList.append(
        {
            "name": d["name"],
            "specializations": sorted(d["specializations"]),
            "accounts": d["accounts"],
            "image": d["image"],
            "postCount": len(pl),
            "totalEngagement": sum(x["engagement"] for x in pl),
            "followers": max((x["followers"] for x in pl), default=0),
            "transcriptCount": sum(1 for x in pl if x.get("transcript")),
        }
    )
doctorList.sort(key=lambda x: (x["totalEngagement"], x["postCount"]), reverse=True)

# ---------- Topic recommendations (clinical only) ----------
tstats = defaultdict(lambda: {"posts": 0, "engagement": 0, "views": 0, "examples": []})
for p in posts:
    t = p["topicAudio"] or "Unclassified"
    s = tstats[t]
    s["posts"] += 1
    s["engagement"] += p["engagement"]
    s["views"] += p["views"]
    if p["caption"] and len(s["examples"]) < 2:
        s["examples"].append(p["caption"][:120])
recs = []
for t, s in tstats.items():
    if not is_clinical(t):
        continue
    recs.append(
        {
            "topic": t,
            "posts": s["posts"],
            "engagement": s["engagement"],
            "views": s["views"],
            "score": s["engagement"] + int(0.2 * s["views"]),
            "suggestion": "قدّم هذا الموضوع بصيغة (معلومة سريعة + حالة واقعية + توصية عملية).",
            "examples": s["examples"],
        }
    )
recs.sort(key=lambda x: x["score"], reverse=True)

# ---------- Coverage ----------
linked_codes = set()
for p in posts:
    if p.get("transcript"):
        sc = shortcode(p["transcript"].get("videoUrl", ""))
        if sc:
            linked_codes.add(sc)
unmatched = [t for t in transcripts if t["shortcode"] and t["shortcode"] not in linked_codes]
coverage = {
    "totalPosts": len(posts),
    "totalEntities": len(rawEntities),
    "totalTopicAudit": len(topicAudit),
    "totalTranscripts": len(transcripts),
    "postsWithTranscript": sum(1 for p in posts if p.get("transcript")),
    "transcriptsMatched": len(transcripts) - len(unmatched),
    "transcriptsUnmatched": len(unmatched),
    "doctorsWithoutPosts": sum(1 for d in doctorList if d["postCount"] == 0),
}

DATA = {
    "meta": {
        "title": "نتائج بحث السوق - الجهاز الهضمي",
        "totalPosts": len(posts),
        "totalDoctors": len(doctorList),
    },
    "doctors": doctorList,
    "posts": sorted(posts, key=lambda x: x["rank"]),
    "topicRecommendations": recs,
    "topicAudit": sorted(topicAudit, key=lambda x: x["engagement"], reverse=True),
    "transcripts": transcripts,
    "coverage": coverage,
    "raw": {
        "posts": rawPosts,
        "entities": rawEntities,
        "topics": topicAudit,
        "transcripts": transcripts,
    },
}

out = "const DATA = " + json.dumps(DATA, ensure_ascii=False) + ";\n"
(REPO / "data.js").write_text(out, encoding="utf-8")
print("data.js written:", len(out), "bytes")
print("posts:", len(posts), "| doctors:", len(doctorList), "| entities:", len(rawEntities))
print("coverage:", json.dumps(coverage, ensure_ascii=False))
print("El Sherbiny posts:", sum(1 for p in posts if p["account"] == "Dr. Mohammad El Sherbiny"))
